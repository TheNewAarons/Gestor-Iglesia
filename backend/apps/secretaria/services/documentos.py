from io import BytesIO
from datetime import date

from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, HRFlowable
from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


def _base_doc(buffer, title='Documento'):
    return SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=2.5 * cm,
        leftMargin=2.5 * cm,
        topMargin=2.5 * cm,
        bottomMargin=2.5 * cm,
        title=title,
    )


def _styles():
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(
        name='Titulo',
        fontSize=16,
        fontName='Helvetica-Bold',
        alignment=TA_CENTER,
        spaceAfter=6,
    ))
    styles.add(ParagraphStyle(
        name='Subtitulo',
        fontSize=12,
        fontName='Helvetica',
        alignment=TA_CENTER,
        spaceAfter=4,
    ))
    styles.add(ParagraphStyle(
        name='Cuerpo',
        fontSize=11,
        fontName='Helvetica',
        alignment=TA_JUSTIFY,
        leading=16,
        spaceAfter=8,
    ))
    styles.add(ParagraphStyle(
        name='Firma',
        fontSize=11,
        fontName='Helvetica',
        alignment=TA_CENTER,
        spaceBefore=40,
    ))
    return styles


def generar_carta_membresia(miembro) -> bytes:
    buffer = BytesIO()
    doc = _base_doc(buffer, 'Carta de Membresía')
    styles = _styles()
    hoy = date.today().strftime('%d de %B de %Y')

    story = [
        Paragraph('IGLESIA DEL NAZARENO', styles['Titulo']),
        Paragraph('Carta de Membresía', styles['Subtitulo']),
        HRFlowable(width='100%', thickness=1, color=colors.HexColor('#1E3A5F')),
        Spacer(1, 0.5 * cm),
        Paragraph(f'Fecha: {hoy}', styles['Normal']),
        Spacer(1, 0.5 * cm),
        Paragraph('A quien corresponda:', styles['Cuerpo']),
        Paragraph(
            f'Por medio de la presente, certificamos que <b>{miembro.nombre_completo}</b> '
            f'es miembro activo de nuestra congregación, con fecha de ingreso '
            f'{miembro.fecha_ingreso.strftime("%d de %B de %Y") if miembro.fecha_ingreso else "—"}. '
            f'Durante su permanencia ha demostrado compromiso y fidelidad a los valores '
            f'y principios de nuestra iglesia.',
            styles['Cuerpo'],
        ),
        Paragraph(
            'Se extiende la presente a petición del interesado para los fines que '
            'estime convenientes.',
            styles['Cuerpo'],
        ),
        Spacer(1, 1 * cm),
        Paragraph('Atentamente,', styles['Normal']),
        Paragraph('_______________________________', styles['Firma']),
        Paragraph('Secretaría de la Iglesia', styles['Firma']),
    ]

    doc.build(story)
    return buffer.getvalue()


def generar_carta_recomendacion(miembro, iglesia_destino: str = '') -> bytes:
    buffer = BytesIO()
    doc = _base_doc(buffer, 'Carta de Recomendación')
    styles = _styles()
    hoy = date.today().strftime('%d de %B de %Y')
    destino = iglesia_destino or 'la iglesia de destino'

    story = [
        Paragraph('IGLESIA DEL NAZARENO', styles['Titulo']),
        Paragraph('Carta de Recomendación', styles['Subtitulo']),
        HRFlowable(width='100%', thickness=1, color=colors.HexColor('#1E3A5F')),
        Spacer(1, 0.5 * cm),
        Paragraph(f'Fecha: {hoy}', styles['Normal']),
        Spacer(1, 0.5 * cm),
        Paragraph(f'Estimados hermanos de {destino}:', styles['Cuerpo']),
        Paragraph(
            f'Nos complace recomendar a <b>{miembro.nombre_completo}</b>, quien ha sido '
            f'miembro fiel de nuestra congregación. Durante su tiempo entre nosotros ha '
            f'demostrado un carácter íntegro y un compromiso genuino con la fe cristiana.',
            styles['Cuerpo'],
        ),
        Paragraph(
            'Los invitamos a recibirle con los brazos abiertos como un hermano en la fe. '
            'Confiamos que será una bendición para su congregación.',
            styles['Cuerpo'],
        ),
        Spacer(1, 1 * cm),
        Paragraph('En Cristo,', styles['Normal']),
        Paragraph('_______________________________', styles['Firma']),
        Paragraph('Secretaría de la Iglesia', styles['Firma']),
    ]

    doc.build(story)
    return buffer.getvalue()


def generar_acta_pdf(acta) -> bytes:
    buffer = BytesIO()
    doc = _base_doc(buffer, f'Acta — {acta.titulo}')
    styles = _styles()

    story = [
        Paragraph('IGLESIA DEL NAZARENO', styles['Titulo']),
        Paragraph(f'Acta de Reunión — {acta.get_tipo_display()}', styles['Subtitulo']),
        HRFlowable(width='100%', thickness=1, color=colors.HexColor('#1E3A5F')),
        Spacer(1, 0.4 * cm),
        Paragraph(f'<b>Título:</b> {acta.titulo}', styles['Normal']),
        Paragraph(f'<b>Fecha:</b> {acta.fecha_reunion}', styles['Normal']),
        Paragraph(f'<b>Lugar:</b> {acta.lugar or "—"}', styles['Normal']),
        Paragraph(f'<b>Presidida por:</b> {acta.presidida_por or "—"}', styles['Normal']),
        Paragraph(f'<b>Estado:</b> {acta.get_estado_display()}', styles['Normal']),
        Spacer(1, 0.4 * cm),
    ]

    if acta.asistentes:
        story.append(Paragraph('<b>Asistentes:</b>', styles['Normal']))
        for a in acta.asistentes:
            nombre = a.get('nombre', '') if isinstance(a, dict) else str(a)
            rol = a.get('rol', '') if isinstance(a, dict) else ''
            story.append(Paragraph(f'• {nombre}{" — " + rol if rol else ""}', styles['Normal']))
        story.append(Spacer(1, 0.3 * cm))

    if acta.orden_del_dia:
        story.append(Paragraph('<b>Orden del Día:</b>', styles['Normal']))
        for i, punto in enumerate(acta.orden_del_dia, 1):
            texto = punto.get('punto', str(punto)) if isinstance(punto, dict) else str(punto)
            story.append(Paragraph(f'{i}. {texto}', styles['Normal']))
        story.append(Spacer(1, 0.3 * cm))

    if acta.acuerdos:
        story.append(Paragraph('<b>Acuerdos:</b>', styles['Normal']))
        for acuerdo in acta.acuerdos:
            if isinstance(acuerdo, dict):
                num = acuerdo.get('numero', '')
                desc = acuerdo.get('descripcion', '')
                resp = acuerdo.get('responsable', '')
                story.append(Paragraph(
                    f'<b>#{num}</b> {desc}{" — Responsable: " + resp if resp else ""}',
                    styles['Normal']
                ))
            else:
                story.append(Paragraph(f'• {acuerdo}', styles['Normal']))
        story.append(Spacer(1, 0.3 * cm))

    if acta.observaciones:
        story.append(Paragraph('<b>Observaciones:</b>', styles['Normal']))
        story.append(Paragraph(acta.observaciones, styles['Cuerpo']))

    story += [
        Spacer(1, 1.5 * cm),
        Paragraph('_______________________________', styles['Firma']),
        Paragraph('Secretaría — Firma y Sello', styles['Firma']),
    ]

    doc.build(story)
    return buffer.getvalue()


def exportar_directorio_excel(miembros_qs) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Directorio'

    header_fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=11)
    headers = ['Nombre', 'Ministerio', 'Teléfono', 'Email', 'Clase', 'Rol', 'Activo', 'Ingreso']

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for row, m in enumerate(miembros_qs.select_related('ministry'), 2):
        ws.cell(row=row, column=1, value=m.nombre_completo)
        ws.cell(row=row, column=2, value=m.ministry.nombre if m.ministry else '')
        ws.cell(row=row, column=3, value=m.telefono)
        ws.cell(row=row, column=4, value=m.email)
        ws.cell(row=row, column=5, value=m.get_clase_display() if m.clase else '')
        ws.cell(row=row, column=6, value=m.get_rol_en_ministerio_display() if m.rol_en_ministerio else '')
        ws.cell(row=row, column=7, value='Sí' if m.activo else 'No')
        ws.cell(row=row, column=8, value=str(m.fecha_ingreso) if m.fecha_ingreso else '')

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def exportar_directorio_pdf(miembros_qs) -> bytes:
    buffer = BytesIO()
    doc = _base_doc(buffer, 'Directorio de Miembros')
    styles = _styles()

    story = [
        Paragraph('IGLESIA DEL NAZARENO', styles['Titulo']),
        Paragraph('Directorio de Miembros', styles['Subtitulo']),
        HRFlowable(width='100%', thickness=1, color=colors.HexColor('#1E3A5F')),
        Spacer(1, 0.5 * cm),
    ]

    for m in miembros_qs.select_related('ministry'):
        line = f'<b>{m.nombre_completo}</b>'
        if m.ministry:
            line += f' — {m.ministry.nombre}'
        if m.telefono:
            line += f' | Tel: {m.telefono}'
        if m.email:
            line += f' | {m.email}'
        story.append(Paragraph(line, styles['Normal']))

    doc.build(story)
    return buffer.getvalue()
