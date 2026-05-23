import io
from datetime import date, datetime
from django.utils.timezone import make_aware
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from django.db.models import Sum
from apps.ministerios.models import CajaMinisterio, MovimientoCaja, Ofrenda
from ..models import ConfiguracionFinanzas, MovimientoTesoreria, CuotaFija
from ..selectors import flujo as flujo_selectors
from . import informe as informe_services

MONTHS = ['', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
          'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']

def _fmt(val):
    """Round and format as int if it\'s a whole number."""
    if val is None:
        return 0
    v = round(float(val))
    return v

def _wc(ws, r, c, value, font=None, align=None, number_format=None):
    """Write cell with optional styling."""
    cell = ws.cell(r, c, value)
    if font:
        cell.font = font
    if align:
        cell.alignment = align
    if number_format:
        cell.number_format = number_format
    return cell

def _get_balances_at_date(cutoff):
    """Compute per-ministry balances at a given cutoff date."""
    balances = {}
    cajas = CajaMinisterio.objects.select_related('ministry').filter(ministry__activo=True)
    for caja in cajas:
        ingresos = float(MovimientoCaja.objects.filter(caja=caja, tipo='ingreso', fecha__lt=cutoff).aggregate(t=Sum('monto'))['t'] or 0)
        egresos = float(MovimientoCaja.objects.filter(caja=caja, tipo='egreso', fecha__lt=cutoff).aggregate(t=Sum('monto'))['t'] or 0)
        balances[caja.ministry.slug] = ingresos - egresos
    return balances

def _get_iglesia_balance_at_date(cutoff):
    ingresos = float(MovimientoTesoreria.objects.filter(tipo__startswith='ingreso_', fecha__lt=cutoff).aggregate(t=Sum('monto'))['t'] or 0)
    egresos = float(MovimientoTesoreria.objects.filter(tipo__startswith='egreso_', fecha__lt=cutoff).aggregate(t=Sum('monto'))['t'] or 0)
    return ingresos - egresos

def _get_month_movements(anio, mes):
    """Get ingresos/egresos for the specified month."""
    movs = MovimientoCaja.objects.filter(fecha__year=anio, fecha__month=mes).select_related('caja__ministry')
    ingresos = {}
    egresos = {}
    for m in movs:
        slug = m.caja.ministry.slug
        val = float(m.monto)
        if m.tipo == 'ingreso':
            ingresos[slug] = ingresos.get(slug, 0) + val
        else:
            egresos[slug] = egresos.get(slug, 0) + val
    return ingresos, egresos

def _get_tes_movements(anio, mes):
    tes_movs = MovimientoTesoreria.objects.filter(fecha__year=anio, fecha__month=mes)
    tes_por_tipo = {}
    for t in tes_movs:
        tes_por_tipo[t.tipo] = tes_por_tipo.get(t.tipo, 0) + float(t.monto)
    return tes_por_tipo

def _get_cuotas():
    cuotas = {}
    for c in CuotaFija.objects.select_related('ministry').all():
        slug = c.ministry.slug
        if slug not in cuotas:
            cuotas[slug] = {}
        cuotas[slug][c.tipo] = float(c.monto)
    return cuotas

def exportar_informe_excel(anio, mes, usuario):
    config = ConfiguracionFinanzas.obtener()
    informe = flujo_selectors.obtener_informe(anio, mes)
    if not informe:
        informe = informe_services.generar_informe(anio, mes, usuario)
    d = informe.datos

    if mes == 1:
        mes_ant = 12
        anio_ant = anio - 1
    else:
        mes_ant = mes - 1
        anio_ant = anio

    mes_nombre = MONTHS[mes]
    mes_ant_nombre = MONTHS[mes_ant].upper()

    cutoff = make_aware(datetime(anio, mes, 1))
    mes_cutoff = make_aware(datetime(anio_ant, mes_ant, 1))

    # Balances at start of month
    ministry_balances = _get_balances_at_date(cutoff)
    iglesia_balance = _get_iglesia_balance_at_date(cutoff)

    principales = ['dni', 'jni', 'mni']
    all_slugs = list(ministry_balances.keys())
    secondary_slugs = [s for s in all_slugs if s not in principales]

    saldo_otros = sum(ministry_balances.get(s, 0) for s in secondary_slugs)
    saldo_dni = ministry_balances.get('dni', 0)
    saldo_jni = ministry_balances.get('jni', 0)
    saldo_mni = ministry_balances.get('mni', 0)
    saldo_total = iglesia_balance + saldo_otros + saldo_dni + saldo_jni + saldo_mni

    # This month movements
    mes_ingresos, mes_egresos = _get_month_movements(anio, mes)
    tes = _get_tes_movements(anio, mes)
    cuotas = _get_cuotas()

    def _ti(key):
        return tes.get(key, 0)

    def _cq(slug, tipo):
        return cuotas.get(slug, {}).get(tipo, 0)

    ofrendas_total = sum(float(o.monto) for o in Ofrenda.objects.filter(fecha__year=anio, fecha__month=mes))
    mni_data = flujo_selectors.ofrendas_por_categoria_mni(mes, anio)

    # Totals
    total_diezmos = _ti('ingreso_diezmo')
    total_especiales = _ti('ingreso_especial')

    ingresos_iglesia = ofrendas_total + total_diezmos + total_especiales
    egresos_iglesia_raw = (
        _ti('egreso_sosten_pastoral') + _ti('egreso_beneficios_pastorales') +
        _ti('egreso_varios_iglesia') + _ti('egreso_fondo_contingencia')
    )
    pres_distrital = round(saldo_total * float(config.pres_distrital_pct) / 100, 2)
    pres_educacional = round(saldo_total * float(config.pres_educacional_pct) / 100, 2)
    pres_evangelismo = round(saldo_total * float(config.pres_evangelismo_pct) / 100, 2)
    coce_cuota = _cq('mni', 'coce') + _cq('dni', 'coce') + _cq('jni', 'coce')
    jubilacion = float(config.jubilacion_monto)
    contingencia = _ti('egreso_fondo_contingencia')

    total_egresos_iglesia = (
        egresos_iglesia_raw + pres_distrital + pres_educacional +
        pres_evangelismo + jubilacion + coce_cuota + _ti('egreso_fondo_contingencia')
    )

    # Secondary ministries
    ingresos_sec = {s: mes_ingresos.get(s, 0) + float(sum(
        o.monto for o in Ofrenda.objects.filter(fecha__year=anio, fecha__month=mes, ministry__slug=s)
    )) for s in secondary_slugs}
    egresos_sec = {s: mes_egresos.get(s, 0) for s in secondary_slugs}
    total_ingresos_sec = sum(ingresos_sec.values())
    total_egresos_sec = sum(egresos_sec.values())

    # DNI, JNI, MNI
    ingresos_dni = mes_ingresos.get('dni', 0) + float(sum(o.monto for o in Ofrenda.objects.filter(fecha__year=anio, fecha__month=mes, ministry__slug='dni')))
    egresos_dni = mes_egresos.get('dni', 0)
    cuota_dni = _cq('dni', 'cuota_distrital')
    coce_dni = _cq('dni', 'coce')

    ingresos_jni = mes_ingresos.get('jni', 0) + float(sum(o.monto for o in Ofrenda.objects.filter(fecha__year=anio, fecha__month=mes, ministry__slug='jni')))
    egresos_jni = mes_egresos.get('jni', 0)
    cuota_jni = _cq('jni', 'cuota_distrital')
    coce_jni = _cq('jni', 'coce')

    ingresos_mni = mes_ingresos.get('mni', 0) + mni_data['total']
    egresos_mni = mes_egresos.get('mni', 0)
    cuota_mni = _cq('mni', 'cuota_distrital')
    coce_mni = _cq('mni', 'coce')

    total_ingresos_mes = ingresos_iglesia + total_ingresos_sec + ingresos_dni + ingresos_jni + ingresos_mni
    total_egresos_mes = (
        total_egresos_iglesia + total_egresos_sec +
        cuota_dni + coce_dni + cuota_jni + coce_jni +
        cuota_mni + coce_mni +
        mni_data['caja_alabastro'] + mni_data['fem']
    )

    saldo_final = saldo_total + total_ingresos_mes - total_egresos_mes

    # Per-category saldos end of month
    iglesia_final = iglesia_balance + ingresos_iglesia - total_egresos_iglesia
    dni_final = saldo_dni + ingresos_dni - (egresos_dni + cuota_dni + coce_dni)
    jni_final = saldo_jni + ingresos_jni - (egresos_jni + cuota_jni + coce_jni)
    mni_final = saldo_mni + ingresos_mni - (egresos_mni + cuota_mni + coce_mni + mni_data['caja_alabastro'] + mni_data['fem'])
    otros_final = saldo_otros + total_ingresos_sec - total_egresos_sec

    # ===== STYLES =====
    b = Font(bold=True, size=10)
    b11 = Font(bold=True, size=11)
    b14 = Font(bold=True, size=14)
    n = Font(size=10)
    left = Alignment(horizontal='left')
    center = Alignment(horizontal='center')
    right = Alignment(horizontal='right')

    # ===== WORKBOOK =====
    wb = Workbook()

    # ==========================================
    # SHEET 1: Informe de Finanzas
    # ==========================================
    ws1 = wb.active
    ws1.title = 'Informe de Finanzas'

    ws1.column_dimensions['A'].width = 9
    ws1.column_dimensions['B'].width = 27
    ws1.column_dimensions['C'].width = 14
    ws1.column_dimensions['D'].width = 14
    ws1.column_dimensions['E'].width = 4
    ws1.column_dimensions['F'].width = 5
    ws1.column_dimensions['G'].width = 30
    ws1.column_dimensions['H'].width = 14
    ws1.column_dimensions['I'].width = 14
    ws1.column_dimensions['K'].width = 35

    iglesia_name = config.nombre_iglesia
    distrito = config.nombre_distrito

    # Header
    ws1.merge_cells('A1:I1')
    _wc(ws1, 1, 1, f'INFORME DE FINANZAS {iglesia_name.upper()}', b14, center)
    ws1.merge_cells('A2:G2')
    _wc(ws1, 2, 1, distrito, b11)
    _wc(ws1, 2, 8, mes_nombre.upper(), b11)
    _wc(ws1, 2, 9, anio, b11)

    # Row 3 headers
    ws1.merge_cells('A3:D3')
    _wc(ws1, 3, 1, f'SALDO MES PASADO: {mes_ant_nombre}', b)
    ws1.merge_cells('E3:I3')
    _wc(ws1, 3, 5, 'DINERO EGRESADO:', b)

    # Left: Saldos
    _wc(ws1, 4, 1, 'OTROS MINISTERIOS', b)
    _wc(ws1, 4, 3, _fmt(saldo_otros), n, right, '#,##0')
    _wc(ws1, 5, 1, 'IGLESIA', b)
    _wc(ws1, 5, 3, _fmt(iglesia_balance), n, right, '#,##0')
    _wc(ws1, 6, 1, 'D.N.I.', b)
    _wc(ws1, 6, 3, _fmt(saldo_dni), n, right, '#,##0')
    _wc(ws1, 7, 1, 'J.N.I.', b)
    _wc(ws1, 7, 3, _fmt(saldo_jni), n, right, '#,##0')
    _wc(ws1, 8, 1, 'M.N.I.', b)
    _wc(ws1, 8, 3, _fmt(saldo_mni), n, right, '#,##0')
    _wc(ws1, 9, 1, 'Total Saldo mes Pasado', b)
    _wc(ws1, 9, 4, _fmt(saldo_total), b, right, '#,##0')

    # Right: Egresos iglesia
    ws1.merge_cells('E4:G4')
    _wc(ws1, 4, 5, 'POR IGLESIA LOCAL:', b)

    egresos_list = [
        (1, 'SOSTEN PASTORAL', _ti('egreso_sosten_pastoral')),
        (2, 'VARIOS IGLESIA LOCAL', _ti('egreso_varios_iglesia')),
        (3, 'PRES. DISTRITAL 10%', pres_distrital),
        (4, 'PRES. EDUCACIONAL 3%', pres_educacional),
        (5, 'PRES. EVANGELISMO 2%', pres_evangelismo),
        (6, 'CENTRO DE MINISTERIOS', cuota_mni + cuota_dni + cuota_jni),
        (7, 'JUBILACION', jubilacion),
        (8, 'COCE', 0),
        (9, 'FONDO ASM. GENERAL', 0),
        (10, 'FONDO DE CONTINGENCIA', _ti('egreso_fondo_contingencia')),
    ]

    for i, (num, nombre, val) in enumerate(egresos_list):
        r = 5 + i
        _wc(ws1, r, 5, 'S.', n)
        _wc(ws1, r, 6, num, n)
        _wc(ws1, r, 7, nombre, n)
        _wc(ws1, r, 8, _fmt(val), n, right, '#,##0')

    # Left: Entradas iglesia
    ws1.merge_cells('A10:B10')
    _wc(ws1, 10, 1, 'POR IGLESIA LOCAL:', b)

    entradas_list = [
        (1, 'OFRENDAS', ofrendas_total),
        (2, 'DIEZMOS', total_diezmos),
        (3, 'ESPECIALES', total_especiales),
    ]
    for i, (num, nombre, val) in enumerate(entradas_list):
        r = 11 + i
        _wc(ws1, r, 1, num, n)
        _wc(ws1, r, 2, nombre, n)
        _wc(ws1, r, 3, _fmt(val), n, right, '#,##0')

    _wc(ws1, 14, 3, _fmt(ingresos_iglesia), n, right, '#,##0')
    _wc(ws1, 15, 2, 'Total entradas iglesia', b)
    _wc(ws1, 15, 4, _fmt(ingresos_iglesia), b, right, '#,##0')

    # Right: Total gastos iglesia row
    _wc(ws1, 15, 7, 'Total gastos iglesia', b)
    _wc(ws1, 15, 9, _fmt(total_egresos_iglesia), b, right, '#,##0')

    # Left: Otros ministerios
    _wc(ws1, 16, 1, 'POR OTROS MINISTERIOS:', b)

    # Secondary ministry list (limited to what fits)
    sec_names = {
        'compasion': 'COMPASIÓN', 'adulto-mayor': 'ADULTO MAYOR',
        'mam': 'MAM', 'oracion': 'ORACIÓN', 'proyectos': 'PROYECTOS',
        'evangelismo': 'EVANGELISMO', 'alabanza': 'ALABANZA',
        'nazakids': 'NAZAKIDS', 'comunicaciones': 'COMUNIC.',
        'danza': 'DANZA', 'teatro': 'TEATRO', 'vid': 'EXPLO',
        'nazartes': 'NAZARTES',
    }

    sec_items = []
    sec_idx = 5
    for s in secondary_slugs:
        name = sec_names.get(s, s.upper().replace('-', ' '))
        sec_items.append((sec_idx, name, ingresos_sec.get(s, 0)))
        sec_idx += 1

    for i, (num, nombre, val) in enumerate(sec_items):
        r = 17 + i
        _wc(ws1, r, 1, num, n)
        _wc(ws1, r, 2, nombre, n)
        _wc(ws1, r, 3, _fmt(val), n, right, '#,##0')

    sec_end = 17 + len(sec_items)
    _wc(ws1, sec_end, 2, 'Total otros ministerios', b)
    _wc(ws1, sec_end, 4, _fmt(total_ingresos_sec), b, right, '#,##0')

    # Right: Total gastos otros ministerios header
    _wc(ws1, 16, 7, 'Total Gastos otros ministerios', b)

    # Right: Secondary egresos
    eg_sec_items = []
    eg_idx = 11 + len(egresos_list)  # continue numbering
    for s in secondary_slugs:
        name = sec_names.get(s, s.upper().replace('-', ' '))
        val = egresos_sec.get(s, 0)
        eg_sec_items.append((eg_idx, name, val))
        eg_idx += 1

    for i, (num, nombre, val) in enumerate(eg_sec_items):
        r = 17 + i
        _wc(ws1, r, 5, 'S.', n)
        _wc(ws1, r, 6, num, n)
        _wc(ws1, r, 7, nombre, n)
        _wc(ws1, r, 8, _fmt(val), n, right, '#,##0')

    eg_sec_end = 17 + len(eg_sec_items)
    _wc(ws1, eg_sec_end, 7, 'Total otros ministerios', b)
    _wc(ws1, eg_sec_end, 9, _fmt(total_egresos_sec), b, right, '#,##0')

    if eg_sec_end > sec_end:
        max_end = eg_sec_end
    else:
        max_end = sec_end

    row_cur = max_end + 1

    # DNI section
    _wc(ws1, row_cur, 1, 'POR D.N.I.:', b)
    _wc(ws1, row_cur, 5, 'POR D.N.I.:', b)

    eg_num = eg_idx + 1
    row_cur += 1
    _wc(ws1, row_cur, 1, eg_num - 1 if eg_num > 1 else 19, n)
    _wc(ws1, row_cur, 2, 'OFRENDAS', n)
    _wc(ws1, row_cur, 3, _fmt(ingresos_dni), n, right, '#,##0')

    _wc(ws1, row_cur, 5, 'S.', n)
    _wc(ws1, row_cur, 6, eg_num, n)
    _wc(ws1, row_cur, 7, 'GASTOS LOCALES', n)
    _wc(ws1, row_cur, 8, _fmt(egresos_dni), n, right, '#,##0')
    eg_num += 1

    row_cur += 1
    _wc(ws1, row_cur, 1, 'Total Entradas D.N.I.', b)
    _wc(ws1, row_cur, 4, _fmt(ingresos_dni), b, right, '#,##0')

    _wc(ws1, row_cur, 5, 'S.', n)
    _wc(ws1, row_cur, 6, eg_num, n)
    _wc(ws1, row_cur, 7, 'CUOTA DISTRITAL', n)
    _wc(ws1, row_cur, 8, _fmt(cuota_dni), n, right, '#,##0')
    eg_num += 1

    row_cur += 1
    _wc(ws1, row_cur, 5, 'S.', n)
    _wc(ws1, row_cur, 6, eg_num, n)
    _wc(ws1, row_cur, 7, 'COCE', n)
    _wc(ws1, row_cur, 8, _fmt(coce_dni), n, right, '#,##0')
    eg_num += 1

    row_cur += 1
    _wc(ws1, row_cur, 5, 'Total Gastos del D.N.I.', b)
    gastos_dni_total = egresos_dni + cuota_dni + coce_dni
    _wc(ws1, row_cur, 9, _fmt(gastos_dni_total), b, right, '#,##0')

    # JNI section
    row_cur += 1
    _wc(ws1, row_cur, 1, 'POR J.N.I.:', b)
    _wc(ws1, row_cur, 5, 'POR J.N.I.:', b)

    row_cur += 1
    _wc(ws1, row_cur, 1, eg_num, n)
    _wc(ws1, row_cur, 2, 'OFRENDAS', n)
    _wc(ws1, row_cur, 3, _fmt(ingresos_jni), n, right, '#,##0')

    _wc(ws1, row_cur, 5, 'S.', n)
    _wc(ws1, row_cur, 6, eg_num + 1, n)
    _wc(ws1, row_cur, 7, 'GASTOS LOCALES', n)
    _wc(ws1, row_cur, 8, _fmt(egresos_jni), n, right, '#,##0')
    eg_num += 2

    row_cur += 1
    _wc(ws1, row_cur, 1, 'Total Entradas J.N.I.', b)
    _wc(ws1, row_cur, 4, _fmt(ingresos_jni), b, right, '#,##0')

    _wc(ws1, row_cur, 5, 'S.', n)
    _wc(ws1, row_cur, 6, eg_num, n)
    _wc(ws1, row_cur, 7, 'CUOTA DISTRITAL', n)
    _wc(ws1, row_cur, 8, _fmt(cuota_jni), n, right, '#,##0')
    eg_num += 1

    row_cur += 1
    _wc(ws1, row_cur, 5, 'S.', n)
    _wc(ws1, row_cur, 6, eg_num, n)
    _wc(ws1, row_cur, 7, 'COCE', n)
    _wc(ws1, row_cur, 8, _fmt(coce_jni), n, right, '#,##0')
    eg_num += 1

    row_cur += 1
    _wc(ws1, row_cur, 5, 'Total Gastos de J.N.I.', b)
    gastos_jni_total = egresos_jni + cuota_jni + coce_jni
    _wc(ws1, row_cur, 9, _fmt(gastos_jni_total), b, right, '#,##0')

    # MNI section
    row_cur += 1
    _wc(ws1, row_cur, 1, 'POR M.N.I.:', b)
    _wc(ws1, row_cur, 5, 'POR M.N.I.:', b)

    # Left: MNI ofrendas detail
    mni_left_items = [
        ('OFRENDAS', mni_data['ofrenda_general']),
        ('ORACION Y AYUNO', mni_data['oracion_ayuno']),
        ('RESURRECCION', mni_data.get('dip', 0)),
        ('ACCION DE GRACIAS', mni_data['accion_gracias']),
        ('CAJA DE ALABASTRO', mni_data['caja_alabastro']),
        ('TMM RADIO', mni_data.get('otros', 0)),
        ('PLAN MEDICO', 0),
        ('SEMINARIOS', 0),
        ('FEM', mni_data['fem']),
        ('ESPECIAL', 0),
    ]

    for i, (nombre, val) in enumerate(mni_left_items):
        row_cur += 1
        _wc(ws1, row_cur, 1, eg_num + i, n)
        _wc(ws1, row_cur, 2, nombre, n)
        _wc(ws1, row_cur, 3, _fmt(val), n, right, '#,##0')

    row_cur += 1
    _wc(ws1, row_cur, 1, 'Total Entradas de M.N.I.', b)
    _wc(ws1, row_cur, 4, _fmt(ingresos_mni), b, right, '#,##0')

    # Right: MNI egresos
    mni_base_row = row_cur - len(mni_left_items)
    mni_right_items = [
        ('GASTOS LOCALES', egresos_mni),
        ('CUOTA DISTRITAL', cuota_mni),
        ('COCE', coce_mni),
        ('RESURRECCION', mni_data.get('dip', 0)),
        ('CAJA DE ALABASTRO', mni_data['caja_alabastro']),
        ('ACCION DE GRACIAS', 0),
        ('FEM', mni_data['fem']),
    ]

    for i, (nombre, val) in enumerate(mni_right_items):
        r = mni_base_row + i
        _wc(ws1, r, 5, 'S.', n)
        _wc(ws1, r, 6, eg_num + i, n)
        _wc(ws1, r, 7, nombre, n)
        _wc(ws1, r, 8, _fmt(val), n, right, '#,##0')

    mni_right_end = mni_base_row + len(mni_right_items)
    _wc(ws1, mni_right_end, 5, 'Total Gastos de M.N.I.', b)
    gastos_mni_total = egresos_mni + cuota_mni + coce_mni + mni_data.get('dip', 0) + mni_data['caja_alabastro'] + mni_data['fem']
    _wc(ws1, mni_right_end, 9, _fmt(gastos_mni_total), b, right, '#,##0')

    # Bottom totals
    row_cur += 1
    _wc(ws1, row_cur, 5, 'TOTAL GASTOS DEL MES', b)
    _wc(ws1, row_cur, 9, _fmt(total_egresos_mes), b, right, '#,##0')

    row_cur += 1
    _wc(ws1, row_cur, 1, 'TOTAL ENTRADAS DEL MES', b)
    _wc(ws1, row_cur, 4, _fmt(total_ingresos_mes), b, right, '#,##0')
    _wc(ws1, row_cur, 5, 'Otros ministerios', n)
    _wc(ws1, row_cur, 8, _fmt(otros_final), n, right, '#,##0')

    row_cur += 1
    _wc(ws1, row_cur, 1, 'Saldo mes Pasado', n)
    _wc(ws1, row_cur, 3, _fmt(saldo_total), n, right, '#,##0')
    _wc(ws1, row_cur, 5, 'Iglesia', n)
    _wc(ws1, row_cur, 8, _fmt(iglesia_final), n, right, '#,##0')

    row_cur += 1
    _wc(ws1, row_cur, 1, 'Total Ingresos del mes', n)
    _wc(ws1, row_cur, 3, _fmt(total_ingresos_mes), n, right, '#,##0')
    _wc(ws1, row_cur, 5, 'D.N.I.', n)
    _wc(ws1, row_cur, 8, _fmt(dni_final), n, right, '#,##0')

    row_cur += 1
    _wc(ws1, row_cur, 1, 'Total Ingr. Y Saldo', b)
    _wc(ws1, row_cur, 3, _fmt(saldo_total + total_ingresos_mes), b, right, '#,##0')
    _wc(ws1, row_cur, 5, 'J.N.I.', n)
    _wc(ws1, row_cur, 8, _fmt(jni_final), n, right, '#,##0')

    row_cur += 1
    _wc(ws1, row_cur, 1, 'Total Gastos del mes', n)
    _wc(ws1, row_cur, 3, _fmt(total_egresos_mes), n, right, '#,##0')
    _wc(ws1, row_cur, 5, 'M.N.I.', n)
    _wc(ws1, row_cur, 8, _fmt(mni_final), n, right, '#,##0')

    row_cur += 1
    _wc(ws1, row_cur, 1, 'SALDO FIN DEL MES', b)
    _wc(ws1, row_cur, 4, _fmt(saldo_final), b, right, '#,##0')
    _wc(ws1, row_cur, 5, 'TOTAL DE SALDO EN CAJA:', b)
    _wc(ws1, row_cur, 9, _fmt(iglesia_final + otros_final + dni_final + jni_final + mni_final), b, right, '#,##0')

    # ==========================================
    # SHEET 2: Informe Mensual de la Tesorería
    # ==========================================
    ws2 = wb.create_sheet('Informe Mensual Tesorería')
    ws2.column_dimensions['A'].width = 12
    ws2.column_dimensions['B'].width = 40
    ws2.column_dimensions['C'].width = 12
    ws2.column_dimensions['D'].width = 12
    ws2.column_dimensions['E'].width = 12
    ws2.column_dimensions['F'].width = 12
    ws2.column_dimensions['G'].width = 10
    ws2.column_dimensions['H'].width = 14
    ws2.column_dimensions['I'].width = 10
    ws2.column_dimensions['J'].width = 14

    ws2.merge_cells('A1:J1')
    _wc(ws2, 1, 1, f'INFORME MENSUAL DE LA TESORERIA {iglesia_name.upper()} {config.ciudad.upper()}', b14, center)
    ws2.merge_cells('A2:E2')
    _wc(ws2, 2, 1, 'FECHA INCLUIDA:', b)
    _wc(ws2, 2, 6, mes_nombre.upper(), b)
    _wc(ws2, 2, 7, anio, b)

    _wc(ws2, 3, 2, 'SALDO', b)

    _wc(ws2, 4, 2, 'IGLESIA', n)
    _wc(ws2, 4, 9, '$', n)
    _wc(ws2, 4, 10, _fmt(iglesia_balance), n, right, '#,##0')

    _wc(ws2, 5, 2, 'ENTRADAS IGLESIA', b)

    _wc(ws2, 6, 2, 'OFRENDAS', n)
    _wc(ws2, 6, 7, '$', n)
    _wc(ws2, 6, 8, _fmt(ofrendas_total), n, right, '#,##0')

    _wc(ws2, 7, 2, 'DIEZMOS', n)
    _wc(ws2, 7, 7, '$', n)
    _wc(ws2, 7, 8, _fmt(total_diezmos), n, right, '#,##0')

    _wc(ws2, 8, 2, 'ESPECIALES', n)
    _wc(ws2, 8, 7, '$', n)
    _wc(ws2, 8, 8, _fmt(total_especiales), n, right, '#,##0')

    _wc(ws2, 9, 10, _fmt(ingresos_iglesia), n, right, '#,##0')

    _wc(ws2, 10, 2, 'TOTAL ENTRADA IGLESIA', b)
    _wc(ws2, 10, 10, _fmt(ingresos_iglesia), b, right, '#,##0')

    _wc(ws2, 11, 2, 'SUMA SALDO Y ENTRADA IGLESIA', b)
    _wc(ws2, 11, 10, _fmt(iglesia_balance + ingresos_iglesia), b, right, '#,##0')

    _wc(ws2, 12, 2, 'EGRESOS IGLESIA', b)

    egresos_detalle = [
        ('sueldo pastoral', _ti('egreso_sosten_pastoral')),
        ('Mes por año en cuotas mensuales', 0),
        ('LUZ', 0),
        ('AGUA', 0),
        ('Internet', 0),
        ('Zoom', 0),
        ('Celular', 0),
        ('2 Mesa Peglables', 0),
        ('Thermo Baño', 0),
        ('ofrenda amor', 0),
        ('4% AHORRO', _ti('ingreso_ahorro') * 0.04 if _ti('ingreso_ahorro') else 0),
    ]

    for i, (nombre, val) in enumerate(egresos_detalle):
        _wc(ws2, 13 + i, 2, nombre, n)
        _wc(ws2, 13 + i, 7, '$', n)
        _wc(ws2, 13 + i, 8, _fmt(val) if val else '', n, right, '#,##0')

    # Subtotal varios iglesia
    iglesia_detail_end = 13 + len(egresos_detalle)
    _wc(ws2, iglesia_detail_end, 10, _fmt(egresos_iglesia_raw), n, right, '#,##0')

    # Fijos
    fijos = [
        ('prs. Distrital 10%', pres_distrital),
        ('prs. Educacional 3%', pres_educacional),
        ('prs. Evangelismo 2%', pres_evangelismo),
        ('centro de ministerios', coce_cuota),
        ('jubilacion', jubilacion),
        ('campamento coce', 0),
        ('fondo asamblea general', 0),
        ('fondo contingencia', _ti('egreso_fondo_contingencia')),
    ]

    for i, (nombre, val) in enumerate(fijos):
        r = iglesia_detail_end + 1 + i
        _wc(ws2, r, 2, nombre, n)
        _wc(ws2, r, 5, 'Fijo', n)
        _wc(ws2, r, 7, '$', n)
        _wc(ws2, r, 8, _fmt(val) if val else '', n, right, '#,##0')

    fijos_end = iglesia_detail_end + 1 + len(fijos)
    _wc(ws2, fijos_end, 2, 'EGRESO TOTAL IGLESIA', b)
    _wc(ws2, fijos_end, 10, _fmt(total_egresos_iglesia), b, right, '#,##0')

    _wc(ws2, fijos_end + 1, 2, f'SALDO EN CAJA IGLESIA MES DE', b)
    _wc(ws2, fijos_end + 1, 5, mes_nombre.upper(), b)
    _wc(ws2, fijos_end + 1, 6, anio, b)
    _wc(ws2, fijos_end + 1, 10, _fmt(iglesia_final), b, right, '#,##0')

    _wc(ws2, fijos_end + 2, 2, '', n)  # signature line

    # ==========================================
    # SHEET 3: Comprobante para Depósitos Distritales
    # ==========================================
    ws3 = wb.create_sheet('Comprobante Depósitos')

    ws3.column_dimensions['A'].width = 8
    ws3.column_dimensions['B'].width = 55
    ws3.column_dimensions['C'].width = 10
    ws3.column_dimensions['D'].width = 14
    ws3.column_dimensions['E'].width = 14
    ws3.column_dimensions['F'].width = 14
    ws3.column_dimensions['G'].width = 10
    ws3.column_dimensions['H'].width = 14
    ws3.column_dimensions['I'].width = 14

    _wc(ws3, 1, 2, f'{iglesia_name}-{distrito}', b11)
    ws3.merge_cells('B2:I2')
    _wc(ws3, 2, 2, 'TESORERIA DISTRITAL', Font(bold=True, size=13), center)

    _wc(ws3, 3, 2, 'COMPROBANTE PARA DEPOSITOS DISTRITALES', b)

    _wc(ws3, 5, 4, 'CORRESPONDIENTE DEL MES:', n)
    _wc(ws3, 5, 6, mes_nombre.upper(), b)
    _wc(ws3, 5, 8, anio, b)

    # District deposit items
    depositos = [
        ('A1', 'ADMINISTRACION DISTRITAL (10% diezmos y ofrendas)', pres_distrital),
        ('C2', 'MINISTERIO ESCUELA DOMINICAL (Cuota $ {:.0f})'.format(cuota_dni), cuota_dni),
        ('C3', 'JUVENTUD NAZARENA INTERNACIONAL (Cuota $ {:.0f})'.format(cuota_jni), cuota_jni),
        ('C4', 'MISIONES NAZARENAS INTERNACIONAL (Cuota $ {:.0f})'.format(cuota_mni), cuota_mni),
        ('C5', 'COMITÉ ORGANIZADOR DE CAMP.Y ENCUENTROS', 0),
    ]
    row = 7
    for codigo, nombre, val in depositos:
        _wc(ws3, row, 1, codigo, n)
        _wc(ws3, row, 2, nombre, n)
        _wc(ws3, row, 7, '$', n)
        _wc(ws3, row, 8, _fmt(val), n, right, '#,##0')
        row += 1

    _wc(ws3, row, 2, '(Cuota $ 200, IGLESIA, MED, JNI, MNI: TOTAL $ 800)', n)
    _wc(ws3, row, 7, '$', n)
    _wc(ws3, row, 8, _fmt(800), n, right, '#,##0')
    row += 1

    _wc(ws3, row, 1, 'C6', n)
    _wc(ws3, row, 2, 'EDUCACION TEOLOGICA (3% Diezmos y Ofrendas)', n)
    _wc(ws3, row, 7, '$', n)
    _wc(ws3, row, 8, _fmt(pres_educacional), n, right, '#,##0')
    row += 1

    _wc(ws3, row, 1, 'C7', n)
    _wc(ws3, row, 2, 'EVANGELISMO (2% Diezmos y Ofrendas)', n)
    _wc(ws3, row, 7, '$', n)
    _wc(ws3, row, 8, _fmt(pres_evangelismo), n, right, '#,##0')
    row += 1

    _wc(ws3, row, 1, 'C8', n)
    _wc(ws3, row, 2, 'PARA CENTRO DE MINISTERIOS Y DISTRITAL (Según cuota)', n)
    _wc(ws3, row, 7, '$', n)
    _wc(ws3, row, 8, _fmt(coce_mni + coce_dni + coce_jni), n, right, '#,##0')
    row += 1

    _wc(ws3, row, 1, 'C9', n)
    _wc(ws3, row, 2, 'FONDO VIAJE ASAMBLEA 2021 PARA LAICO ($ 5.000)', n)
    _wc(ws3, row, 7, '$', n)
    _wc(ws3, row, 8, 0, n, right, '#,##0')
    row += 1

    _wc(ws3, row, 2, 'Fondos Para la Evangelizacion Mundial (Presupuesto General)', b)
    row += 1

    world_items = [
        ('D1', 'ORACION Y AYUNO', mni_data['oracion_ayuno']),
        ('D1', 'RESURRECCION', mni_data.get('dip', 0)),
        ('D1', 'ACCION DE GRACIAS', mni_data['accion_gracias']),
        ('D1', 'OFRENDA DE EVANGELISMO (Sobre mensual FEM)', mni_data['fem']),
    ]
    for codigo, nombre, val in world_items:
        _wc(ws3, row, 1, codigo, n)
        _wc(ws3, row, 2, nombre, n)
        _wc(ws3, row, 7, '$', n)
        _wc(ws3, row, 8, _fmt(val) if val else '', n, right, '#,##0')
        row += 1

    _wc(ws3, row, 2, 'Otras Ofrendas Misioneras', b)
    row += 1

    other_items = [
        ('D2', 'CAJA ALABASTRO', mni_data['caja_alabastro']),
        ('D3', 'TRASMISIONES DE MISION MUNDIAL "TMM"', mni_data.get('otros', 0)),
        ('D4', 'OTRA', 0),
    ]
    for codigo, nombre, val in other_items:
        _wc(ws3, row, 1, codigo, n)
        _wc(ws3, row, 2, nombre, n)
        _wc(ws3, row, 7, '$', n)
        _wc(ws3, row, 8, _fmt(val) if val else '', n, right, '#,##0')
        row += 1

    _wc(ws3, row, 1, 'H7', n)
    _wc(ws3, row, 2, 'JUBILACION PASTORAL ($4.000)', n)
    _wc(ws3, row, 7, '$', n)
    _wc(ws3, row, 8, _fmt(jubilacion) if jubilacion else '', n, right, '#,##0')
    row += 1

    _wc(ws3, row, 1, 'H8', n)
    _wc(ws3, row, 2, 'PARA CONTINGENCIAS (EMERGENCIA) ($3.000)', n)
    _wc(ws3, row, 7, '$', n)
    _wc(ws3, row, 8, _fmt(contingencia) if contingencia else '', n, right, '#,##0')
    row += 1

    # Total
    ws3.cell(row, 6, 'TOTAL').font = b
    ws3.cell(row, 7, '$').font = b
    total_depositos = sum(v for *_, v in depositos + world_items + other_items if isinstance(v, (int, float))) + 800 + pres_educacional + pres_evangelismo + (coce_mni + coce_dni + coce_jni) + jubilacion + contingencia
    _wc(ws3, row, 8, _fmt(total_depositos), b, right, '#,##0')
    row += 2

    # Bank info
    ws3.merge_cells(f'B{row}:H{row}')
    _wc(ws3, row, 2, 'DEPOSITAR EN CUENTA CORRIENTE N° 105-09198-07 EN EL BANCO DE CHILE ARICA A', n)
    row += 1
    ws3.merge_cells(f'B{row}:H{row}')
    _wc(ws3, row, 2, 'NOMBRE DE: IGLESIA DEL NAZARENO EVANGELICA DE CHILE', n)
    row += 1
    ws3.merge_cells(f'B{row}:H{row}')
    _wc(ws3, row, 2, 'ENVIAR COPIA DE ESTE COMPROBANTE JUNTO A COPIA DE DEPOSITO E INFORME', n)
    row += 1
    ws3.merge_cells(f'B{row}:H{row}')
    _wc(ws3, row, 2, 'MENSUAL DE LA TESORERIA DE LA IGLESIA, A LA TESORERA (o entregar a la persona', n)
    row += 1
    _wc(ws3, row, 2, 'encargada).', n)

    row += 1
    _wc(ws3, row, 4, 'GLORIA MORALES ZEGARRA', n)
    row += 1
    _wc(ws3, row, 4, 'Mail: gloria_046@hotmail.com', n)
    row += 1
    _wc(ws3, row, 4, 'Celular 09-1419938', n)

    row += 1
    _wc(ws3, row, 1, 'NOMBRE DEL DEPOSITANTE:', n)
    _wc(ws3, row, 4, '', n)  # empty for manual entry

    row += 1
    ws3.merge_cells(f'A{row}:I{row}')
    _wc(ws3, row, 1, 'NOTA: El informe Mensual de TESORERIA debe ser enviado al Superintendente de Distrito, JUNTO AL INFORME', n)
    row += 1
    _wc(ws3, row, 1, 'MENSUAL DEL PASTOR.', n)

    row += 3
    _wc(ws3, row, 5, 'COLOCAR AQUÍ COPIA DEL RECIBO DE DEPÓSITO', n)
    row += 1
    _wc(ws3, row, 5, 'NUMERO DEPÓSITO', n)

    # ==========================================
    # SHEET 4: Saldo Otros Ministerios
    # ==========================================
    ws4 = wb.create_sheet('Saldo Otros Ministerios')
    ws4.column_dimensions['A'].width = 6
    ws4.column_dimensions['B'].width = 25
    ws4.column_dimensions['C'].width = 12
    ws4.column_dimensions['D'].width = 16
    ws4.column_dimensions['E'].width = 16
    ws4.column_dimensions['F'].width = 16
    ws4.column_dimensions['G'].width = 16

    ws4.merge_cells('C1:G1')
    _wc(ws4, 1, 3, 'INFORME MENSUAL DE LA TESORERIA', b14, center)
    ws4.merge_cells('C2:G2')
    _wc(ws4, 2, 3, iglesia_name, b11, center)
    _wc(ws4, 3, 5, config.ciudad.upper(), b11, center)

    _wc(ws4, 5, 4, 'SALDO OTROS MINISTERIOS', b, center)

    _wc(ws4, 7, 2, 'SALDO ANTERIOR', n)
    _wc(ws4, 7, 5, mes_ant_nombre, n)
    _wc(ws4, 7, 6, anio_ant, n)
    _wc(ws4, 7, 7, _fmt(saldo_otros), n, right, '#,##0')

    _wc(ws4, 8, 2, 'FECHA INCLUIDA', n)

    _wc(ws4, 10, 2, f'SALDOS AL 01 de', n)
    _wc(ws4, 10, 5, mes_nombre.upper(), n)
    _wc(ws4, 10, 6, anio, n)
    _wc(ws4, 10, 7, _fmt(saldo_otros), n, right, '#,##0')

    # Table header
    _wc(ws4, 12, 4, 'SALDO ANTERIOR', b)
    _wc(ws4, 12, 5, 'INGRESOS', b)
    _wc(ws4, 12, 6, 'EGRESOS', b)
    _wc(ws4, 12, 7, 'SALDO ACTUAL', b)
    ws4.merge_cells('H12:I12')

    r = 13
    for s in secondary_slugs:
        name = sec_names.get(s, s.replace('-', ' ').title())
        saldo_ant = ministry_balances.get(s, 0)
        ingresos = ingresos_sec.get(s, 0)
        egresos = egresos_sec.get(s, 0)
        _wc(ws4, r, 2, name, n)
        _wc(ws4, r, 4, _fmt(saldo_ant), n, right, '#,##0')
        _wc(ws4, r, 5, _fmt(ingresos) if ingresos else '', n, right, '#,##0')
        _wc(ws4, r, 6, _fmt(egresos) if egresos else '', n, right, '#,##0')
        _wc(ws4, r, 7, _fmt(saldo_ant + ingresos - egresos), n, right, '#,##0')
        r += 1

    _wc(ws4, r, 2, 'SALDO TOTAL', b)
    _wc(ws4, r, 7, _fmt(otros_final), b, right, '#,##0')

    # ==========================================
    # SHEET 5: Control Semanal
    # ==========================================
    ws5 = wb.create_sheet('Control Semanal')
    ws5.column_dimensions['A'].width = 14
    ws5.column_dimensions['B'].width = 14
    ws5.column_dimensions['C'].width = 14
    ws5.column_dimensions['D'].width = 14
    ws5.column_dimensions['E'].width = 14
    ws5.column_dimensions['F'].width = 14
    ws5.column_dimensions['G'].width = 14
    ws5.column_dimensions['H'].width = 14
    ws5.column_dimensions['I'].width = 14
    ws5.column_dimensions['J'].width = 14
    ws5.column_dimensions['K'].width = 14

    _wc(ws5, 1, 1, 'mes', n)

    _wc(ws5, 2, 1, 'ingresos', b)
    headers1 = ['ofr. Iglesia', 'especial Igl.', 'diezmo', 'D.N.I.', 'JNI', 'MNI', 'FEM', 'ENFASIS']
    for i, h in enumerate(headers1):
        _wc(ws5, 2, 2 + i, h, b)

    weeks = ['semana 1', 'semana 2', 'semana 3', 'semana 4', 'semana 5', 'deposito']
    for i, w in enumerate(weeks):
        _wc(ws5, 3 + i, 1, w, n)

    _wc(ws5, 9, 1, 'TOTAL', b)
    for c in range(2, 10):
        _wc(ws5, 9, c, 0, n, right, '#,##0')

    # Ministry ingresos section
    _wc(ws5, 13, 1, 'ingresos', b)
    ministry_headers = ['COMPASION', 'AD. MAYOR', 'MAM', 'ORACION', 'PROYECTO', 'ALABANZA', 'COMUNIC.', 'EVANG.']
    for i, h in enumerate(ministry_headers):
        _wc(ws5, 13, 2 + i, h, b)

    for i, w in enumerate(weeks):
        _wc(ws5, 14 + i, 1, w, n)

    _wc(ws5, 20, 1, 'TOTAL', b)
    for c in range(2, 10):
        _wc(ws5, 20, c, 0, n, right, '#,##0')

    # Egresos section
    _wc(ws5, 23, 1, 'egresos', b)
    egresos_headers = ['iglesia', 'jni', 'MAM', 'MIEDD', 'PROYECTO', 'COMPASION', 'ADULTO MAYOR', 'MNI', 'EVANGE', 'NAZA KIDD']
    for i, h in enumerate(egresos_headers):
        _wc(ws5, 23, 2 + i, h, b)

    egreso_lines = [
        'sueldo pastoral', 'AFP', 'BONO', 'MES POR AÑO', 'ZOOM', 'LUZ',
        'TELEFONO', 'AGUA', 'asamblea', '', '2% COMPASION', '2% EVANGELISMO', '4% AHORRO',
    ]
    for i, line in enumerate(egreso_lines):
        _wc(ws5, 24 + i, 1, line, n)

    tot_row = 24 + len(egreso_lines) + 7
    _wc(ws5, tot_row, 1, 'TOTAL', b)
    for c in range(2, 12):
        _wc(ws5, tot_row, c, 0, n, right, '#,##0')

    # ===== SAVE =====
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
